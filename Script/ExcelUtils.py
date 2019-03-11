# -*- coding: utf-8 -*-
'''
Transn is pleased to support the open source community by making this file available.
Copyright (C) 2018 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the BSD 3-Clause License (the "License"); you may not use this file except in compliance with the License. You may obtain a copy of the License at
https://opensource.org/licenses/BSD-3-Clause
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific language governing permissions and limitations under the License.

'''


import openpyxl


__all__ = [
    "ExcelDriver"
]

class ExcelDriver(object):

    def __init__(self,path):
        try:
            #self.workbook = xlrd.open_workbook(path)
            self.workbook = openpyxl.load_workbook(path)
        except Exception as e:
            print(e)

    def getWorksheet(self,sheetName):
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            print(e)

    @staticmethod
    def getCellData(iRownum,iColnum,worksheet):
        data = worksheet.cell(iRownum,iColnum).value
        return data

    def setCellData(self,worksheet, cell, value):
        worksheet[cell.column+str(cell.row)] = value

    def saveExcelFile(self,path):
        self.workbook.save(path)

    @staticmethod
    def getSheetRowCount(worksheet):
        #print("max is ",worksheet.max_row)
        #print("min is", worksheet.min_row)
        if worksheet.max_row==worksheet.min_row:
            if worksheet.max_row:
                return 1
            else:
                pass
        return worksheet.max_row-worksheet.min_row
        pass

    @staticmethod
    def getSheetFirstRow(worksheet):
        return worksheet.min_row
        pass

    @staticmethod
    def getSheetLastRow(worksheet):
        return worksheet.max_row
        pass

    @staticmethod
    def getStepBeginRow(caseId, caseIdColumn, worksheet):
        rowCount = ExcelDriver.getSheetRowCount(worksheet)
        for index in range(rowCount):
            test_step_caseId = ExcelDriver.getCellData(index+2, caseIdColumn, worksheet)
            if caseId == test_step_caseId:
                beginRow = index+2
                    #index = index+1
                return beginRow
        if index == rowCount:
                pass

    @staticmethod
    def getStepLastRow(caseId, caseIdColumn, worksheet):
        rowCount = ExcelDriver.getSheetRowCount(worksheet)-1
        for index in range(rowCount,-1,-1):
            test_step_caseId = ExcelDriver.getCellData(index+2, caseIdColumn, worksheet)
            if caseId == test_step_caseId:
                    lastRow = index+2
                    return lastRow
        if index == 0:
                pass

    @staticmethod
    def getCurrentStepRow(stepId, stepColumn, worksheet):
        rowCount = ExcelDriver.getSheetRowCount(worksheet) - 1
        for index in range(rowCount, -1, -1):
            test_step_stepId = ExcelDriver.getCellData(index + 2, stepColumn, worksheet)
            if stepId == test_step_stepId:
                currentRow = index + 2
                return currentRow


if __name__ == "__main__":

    pass
    # workbook = ExcelDriver(constants.FILE_PATH)
    # TestFlowSheet = workbook.getWorksheet(constants.TEST_FLOW_SHEET_NAME)
    # TestStepSheet = workbook.getWorksheet(constants.TEST_STEP_SHEET_NAME)
    # #cell = workSheet.cell(3,7)
    # #resultCell = workSheet.cell(2, constants.col_TestStep_Result)
    # #workbook.setCellData(workSheet,cell,"PASS")
    # print(workbook.getStepBeginRow("pDefaultPage_pYzt",constants.col_Flow_CaseId,TestFlowSheet))
    # print(workbook.getStepBeginRow("Case01", constants.col_TestCase_CaseId, TestStepSheet))
    #workSheet['D2'] = "hELLO"