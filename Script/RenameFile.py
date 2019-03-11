# -*- coding: utf-8 -*-
'''
Transn is pleased to support the open source community by making this file available.
Copyright (C) 2018 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the BSD 3-Clause License (the "License"); you may not use this file except in compliance with the License. You may obtain a copy of the License at
https://opensource.org/licenses/BSD-3-Clause
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific language governing permissions and limitations under the License.

'''



import openpyxl
import os
import re
import shutil
from nt import chdir

class FileUtils:

    def __init__(self,path):
        #try:
            # self.workbook = xlrd.open_workbook(path)
            self.workbook = openpyxl.load_workbook(path)
        #except:

            #raise RuntimeError()

    def getWorkSheet(self,sheetName):
        #try:
            sheet = self.workbook[sheetName]
            return sheet
        #except:
            #raise RuntimeError()

    def renameFile(self,workbookPath,worksheetName):
        workbook = self.workbook(workbookPath)
        workSheet=workbook.getWorkSheet(worksheetName)

        toFileUrl = FileUtils.getCellData(2, 8, workSheet)
        fileName = FileUtils.getCellData(2, 7, workSheet)
        caseId = FileUtils.getCellData(2, 1, workSheet)
        fromLang = FileUtils.getCellData(2, 3, workSheet)
        toLang = FileUtils.getCellData(2, 4, workSheet)
        #print fileName + "===" + toFileUrl

        fileNum = len(os.listdir(u'E:\\_00000\\0304'))

        for index in range(fileNum):
            testPath = u'E:\\_00000\\0304'
            destPath = u'E:\\_00000\\0304\\changed'
            isExists = os.path.exists(destPath)
            if not isExists:
                os.makedirs(destPath)
            rowNum = index + 2
            itemName = os.listdir(testPath)[index]
            caseId = FileUtils.getCellData(rowNum, 1, workSheet)
            fromLang = FileUtils.getCellData(rowNum, 3, workSheet)
            toLang = FileUtils.getCellData(rowNum, 4, workSheet)
            if (re.match(r"^\d+", itemName)):
                print itemName
                newitem = re.sub(r"_(\w+)", "", itemName)
                name = str(caseId) + "_" + fromLang + "_" + toLang + "_" + newitem
                print name
                shutil.copyfile(testPath + "\\" + itemName, destPath + "\\" + name)




    @staticmethod
    def getCellData(iRownum,iColnum,worksheet):
        data = worksheet.cell(iRownum,iColnum).value
        return data





if __name__ == "__main__":

    #pass
    workbook = FileUtils(u"E:\\_测试组文件\\测试资源\\机翻测试.xlsx")
    workSheet = workbook.getWorkSheet(u"复杂文档翻译结果")

    #cell = workSheet.cell(1,0)
    toFileUrl= FileUtils.getCellData(2,8,workSheet)
    fileName = FileUtils.getCellData(2,7,workSheet)
    caseId= FileUtils.getCellData(2, 1, workSheet)
    fromLang = FileUtils.getCellData(2,3,workSheet)
    toLang = FileUtils.getCellData(2,4,workSheet)
    print fileName+"==="+toFileUrl

    # for item in os.listdir(u'E:\\_00000\\0304'):
    #     #print item
    #     if(re.match(r"^\d+_",item)):
    #         #print item
    #         newitem = re.sub(r"_(\w+)","",item)
    #         print str(caseId)+"_"+fromLang+"_"+toLang+"_"+newitem
    fileNum = len(os.listdir(u'E:\\_00000\\0304'))
    for index in range(fileNum):

        testPath = u'E:\\_00000\\0304'
        destPath = u'E:\\_00000\\0304\changed'
        isExists = os.path.exists(destPath)
        if not isExists:
            os.makedirs(destPath)


        #print index
        rowNum=index+2
        itemName = os.listdir(testPath)[index]
        caseId = FileUtils.getCellData(rowNum, 1, workSheet)
        fromLang = FileUtils.getCellData(rowNum, 3, workSheet)
        toLang = FileUtils.getCellData(rowNum, 4, workSheet)
        if (re.match(r"^\d+", itemName)):
            print itemName
            newitem = re.sub(r"_(\w+)", "", itemName)
            name =  str(caseId) + "_" + fromLang + "_" + toLang + "_" + newitem
            print name
            shutil.copyfile(testPath+"\\"+itemName, destPath+"\\"+name)
            #chdir(os.path.dirname(testPath+itemName))
            #os.rename(itemName, name)
    #resultCell = workSheet.cell(2, constants.col_TestStep_Result)
    #workbook.setCellData(workSheet,cell,"PASS")
    #print(workbook.getStepBeginRow("pDefaultPage_pYzt",constants.col_Flow_CaseId,TestFlowSheet))
    #print(workbook.getStepBeginRow("Case01", constants.col_TestCase_CaseId, TestStepSheet))
    #workSheet['D2'] = "hELLO"